# -*- coding: utf-8 -*-
################################################################################
#
#    Cybrosys Technologies Pvt. Ltd.
#
#    Copyright (C) 2024-TODAY Cybrosys Technologies(<https://www.cybrosys.com>).
#    Author: Manasa T P (odoo@cybrosys.com)
#
#    You can modify it under the terms of the GNU AFFERO
#    GENERAL PUBLIC LICENSE (AGPL v3), Version 3.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU AFFERO GENERAL PUBLIC LICENSE (AGPL v3) for more details.
#
#    You should have received a copy of the GNU AFFERO GENERAL PUBLIC LICENSE
#    (AGPL v3) along with this program.
#    If not, see <http://www.gnu.org/licenses/>.
#
################################################################################
import binascii
import tempfile
from openpyxl.reader.excel import load_workbook

from odoo import fields, models, _
from odoo.exceptions import ValidationError


class UserImport(models.TransientModel):
    """Import User with access right"""
    _name = 'user.import'
    _description = 'User Import'

    file = fields.Binary(string="Upload File", help='Upload the file here')

    def import_file(self):
        """ function to import user from xlsx file """
        if self:
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsx",
                                                 delete=False) as temp_file:
                    # Write the base64-decoded content to the temporary file
                    temp_file.write(binascii.a2b_base64(self.file))
                    temp_file.flush()  # Ensure all data is written

                    # Open the workbook using the temporary file
                    book = load_workbook(temp_file.name)
                    sheet = book.active  # Access the active sheet
            except Exception as e:
                raise ValidationError(
                    _(f"Please choose the correct file: {str(e)}"))
            startline = True
            for i in range(1, sheet.max_row+1):
                if startline:
                    startline = False
                else:
                    line = [cell.value for cell in sheet[i]]  # Get values in the row
                    res_lang = self.env['res.lang']
                    res_groups = self.env['res.groups']
                    res_company = self.env['res.company']
                    user_type = [line[4]]
                    invalid_language = [lang for lang in [line[2]] if
                                        not res_lang.search(
                                            [('code', '=', lang),
                                             ('active', '=', True)])]
                    if invalid_language:
                        raise ValidationError(_("Language %s is not active") % (
                            " ".join(invalid_language)))
                    invalid_company = [res for res in [line[3]] if
                                       not res_company.search(
                                           [('name', '=', res)])]
                    if invalid_company:
                        raise ValidationError(_("Company %s not exists") % (
                            " ".join(invalid_company)))
                    invalid_user = [rec for rec in user_type if
                                    not res_groups.search(
                                        [('full_name', '=', rec)])]
                    if invalid_user:
                        raise ValidationError(_("Invalid User Type %s") % (
                            " ".join(invalid_user)))
                    if line[5]:
                        groups = line[5].split(",")
                        invalid_groups = [rec for rec in groups if
                                          not res_groups.search(
                                              [('full_name', '=', rec)])]
                        if invalid_groups:
                            raise ValidationError(_("Invalid groups %s") % (
                                " ".join(invalid_groups)))
                    else:
                        groups = []
                    access_right = res_groups.search(
                        [('full_name', 'in', groups)]).ids
                    tech_settings = line[6].split(',')
                    tech_settings += user_type
                    total_rights = res_groups.search(
                        [('name', 'in', tech_settings)]).ids
                    group_ids = access_right + total_rights
                    if line[0]:
                        self.env['res.users'].create([{
                            'name': line[0],
                            'login': line[1],
                            'lang': line[2],
                            'company_id': self.env['res.company'].search(
                                [('name', '=', line[3])]).id if line[3] else False,
                            'groups_id': group_ids,
                        }])
                    else:
                        raise ValidationError(_('Please Enter the User Name.'))
